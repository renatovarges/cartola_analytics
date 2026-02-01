import openpyxl
import pandas as pd

file_path = r"C:\Users\User\.gemini\antigravity\scratch\cartola_analytics\input\DIVISÃO VOLANTES E MEIAS.xlsx"

# Abrir workbook com openpyxl para ler cores
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("="*80)
print("LENDO CLASSIFICAÇÃO POR COR DE FUNDO")
print("="*80)

# Mapear colunas (baseado na estrutura vista)
# Col A (1) = TIME
# Col C (3) = JOGADOR
# Col E (5) = 1º VOLANTE (verde se for volante 1)
# Col G (7) = 2º VOLANTE (verde se for volante 2)
# Col I (9) = MEIA (verde se for meia)

results = []

for row_idx in range(4, ws.max_row + 1):  # Pular cabeçalho e linha vazia
    time_cell = ws.cell(row=row_idx, column=1)  # Coluna A
    jogador_cell = ws.cell(row=row_idx, column=3)  # Coluna C
    
    time = time_cell.value
    jogador = jogador_cell.value
    
    if not jogador or pd.isna(jogador) or str(jogador).strip() == '':
        continue
    
    # Verificar qual célula está verde (colunas intercaladas!)
    vol1_cell = ws.cell(row=row_idx, column=5)  # Coluna E (1º VOLANTE)
    vol2_cell = ws.cell(row=row_idx, column=7)  # Coluna G (2º VOLANTE)
    meia_cell = ws.cell(row=row_idx, column=9)  # Coluna I (MEIA)
    
    classificacao = None
    
    # Função para verificar se célula tem preenchimento de cor
    def tem_preenchimento(cell):
        try:
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                # Tem preenchimento
                return True
        except:
            pass
        return False
    
    # Verificar qual coluna tem preenchimento (verde)
    if tem_preenchimento(vol1_cell):
        classificacao = "VOLANTE"
    elif tem_preenchimento(vol2_cell):
        classificacao = "VOLANTE"  
    elif tem_preenchimento(meia_cell):
        classificacao = "MEIA"
    
    # Debug: mostrar para primeiras linhas
    if row_idx <= 20 and jogador:
        print(f"\nLinha {row_idx}: {jogador}")
        print(f"  Vol1 preenchimento: {tem_preenchimento(vol1_cell)}")
        print(f"  Vol2 preenchimento: {tem_preenchimento(vol2_cell)}")
        print(f"  Meia preenchimento: {tem_preenchimento(meia_cell)}")
        print(f"  → Classificação: {classificacao}")
    
    if jogador:
        results.append({
            'TIME': time,
            'JOGADOR': jogador,
            'CLASSIFICACAO': classificacao
        })

df_result = pd.DataFrame(results)

print("\n" + "="*80)
print("RESUMO")
print("="*80)
print(f"\nTotal jogadores: {len(df_result)}")

if len(df_result) > 0:
    print(f"\nClassificações:")
    print(df_result['CLASSIFICACAO'].value_counts())

    print("\n" + "="*80)
    print("PRIMEIROS 30 REGISTROS")
    print("="*80)
    print(df_result.head(30))

    # Salvar em CSV para verificação
    output_csv = r"C:\Users\User\.gemini\antigravity\scratch\cartola_analytics\input\classificacao_meias_volantes.csv"
    df_result.to_csv(output_csv, index=False, encoding='utf-8-sig')
    print(f"\n✅ Salvo em: {output_csv}")
